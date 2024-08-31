import disnake
# import pytube
from asyncio import sleep

import disnake.ext.commands
import config
import for_games
from db import *
import datetime
from typing import Optional
from openpyxl import Workbook
from dpyConsole import Console
from disnake.ui import Button, View
from discord_webhook import DiscordWebhook
from disnake import Embed 
from disnake.voice_client import VoiceClient
from disnake.ext import commands 
from disnake.utils import get

import disnake.ext 


#Переменные необходимые в коде, для его сокращения.
intents = disnake.Intents().all() #разрешения
bot = commands.Bot(command_prefix=config.prefix, intents=intents, test_guilds=[1232407034108973186]) #префикс команд и разрешения
my_console = Console(bot)



#логирование сообщений, первая часть кода логирует в файл в более краткой форме, вторая часть логирует в файл и в #log
async def log(message: disnake.Message):
    now = datetime.datetime.now()
    # log_webhook = DiscordWebhook(url='https://discord.com/api/webhooks/1242202771294261429/kch_F1G9r3k9SdQn1LzpOQtr4fSyuc9ZpAYfE_ad5GWPthLVXSCfIh8xhf_CUx8o-DIo')
    log_embed = Embed(
        title='Сообщение',
        description=message.content
    ).set_author(
        name=message.author.global_name,
        url=message.author.avatar.url
    )
    log_channel = await get_info(message.guild.id, 'logChannel')
    log_embed.add_field(name = 'Пользователь и его ID', value = f'{message.author.mention} ( {message.author.id} )', inline = False)
    log_embed.add_field(name = 'Дата', value = f'{now.strftime("%d/%m/%Y")}')
    log_embed.add_field(name = 'Время', value = f'{now.strftime("%H:%M:%S")}', inline=False)
    log_embed.add_field(name = 'Категория и её ID', value = f'{message.channel.category} ( {message.channel.category.id} )', inline=False)
    log_embed.add_field(name = 'Канал и его ID', value = f'{message.channel.mention} ( {message.channel.id} )')
    log_embed.add_field(name = 'Ссылка на сообщение', value = f'{message.jump_url}', inline=False)
    # log_webhook.add_embed(log_embed)
    channel = bot.get_channel(log_channel)
    await channel.send(embed=log_embed)
    # response = log_webhook.execute() 

@bot.event
async def on_message(message: disnake.Message):
    if message.webhook_id is not None:
        return
    await bot.process_commands(message)
    if message.author.bot != True:
        await log(message)
    await db_add_exp(message)


@bot.command()
async def logs(ctx):
    await ctx.send(file=disnake.File(r'log.txt'))

@bot.command()
async def ping(stx):
    book = Workbook()
    sheep = book.active
    mainRow = 1
    for i in range(len(bot.guilds)):
        sheep.cell(row = mainRow, column = 1, value = bot.guilds[i].name)
        for x in range(len(bot.guilds[i].members)):
            sheep.cell(row = mainRow, column = 2, value = bot.guilds[i].members[x].name)
            sheep.cell(row = mainRow, column = 3, value = bot.guilds[i].members[x].global_name)
            if bot.guilds[i].members[x].avatar == None:
                sheep.cell(row = mainRow, column = 4, value = bot.guilds[i].members[x].default_avatar.url)
            else:
                sheep.cell(row = mainRow, column = 4, value = bot.guilds[i].members[x].avatar.url)
            mainRow += 1
    book.save('parser.xlsx')
    book.close
    await stx.send(file = disnake.File(r'parser.xlsx'))

@bot.command()
async def test(ctx):
    await get_info(ctx.message.guild.id, 'logChannel')

@bot.command()
async def setLogChannel(stx):
    await set_log_channel(stx.message)   

bot.load_extensions("cogs")

my_console.start()
bot.run(config.token)